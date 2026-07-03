from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Image as PdfImage
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.config import STORAGE_DIR
from app.services.analysis_service import analyze_merged_data, detect_exam_type, load_merged_data
from app.services.config_service import get_context_config
from app.services.upload_service import get_upload_status


def build_pdf_export(*, academic_year: str, year: str, semester: str, section: str, exam: str) -> BytesIO:
    analysis, uploads, settings = _load_report_data(
        academic_year=academic_year,
        year=year,
        semester=semester,
        section=section,
        exam=exam,
    )
    rows = _subject_rows(analysis, uploads, settings)
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=0.35 * inch,
        leftMargin=0.35 * inch,
        topMargin=0.35 * inch,
        bottomMargin=0.35 * inch,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("ReportTitle", parent=styles["Title"], fontSize=16, leading=20, textColor=colors.HexColor("#0f766e"))
    section_style = ParagraphStyle("SectionTitle", parent=styles["Heading2"], fontSize=12, leading=15, spaceBefore=10, textColor=colors.HexColor("#0f172a"))
    small_style = ParagraphStyle("Small", parent=styles["Normal"], fontSize=8, leading=10)

    story: list[Any] = [_pdf_header(academic_year, year, semester, section, exam, title_style, small_style), Spacer(1, 8)]

    overall = analysis["overall"]
    story.append(
        _table(
            [
                ["Class Strength", "Students Passed", "Students Failed", "Overall Pass %", "Upload Progress"],
                [
                    overall["classStrength"],
                    _all_pass_count(analysis),
                    max(overall["classStrength"] - _all_pass_count(analysis), 0),
                    f"{overall['passPercentage']}%",
                    f"{len(uploads['uploads'])} / {len(settings.get('subjects', []))}",
                ],
            ],
            header_rows=1,
        )
    )

    story.extend([Spacer(1, 8), Paragraph("Report 1 - Overall Section Analysis", section_style)])
    story.append(
        _table(
            [
                ["Name of the Faculty", "Subject Name", "Subject Code", "Total Strength", "Attended", "Absent", "Passed", "Failed", "% of Pass", "% of Fail"],
                *[
                    [
                        row["faculty"],
                        row["subject"],
                        row["subjectCode"],
                        row["strength"],
                        row["attended"],
                        row["absent"],
                        row["passed"],
                        row["failed"],
                        row["passPercentage"],
                        row["failPercentage"],
                    ]
                    for row in rows
                ],
            ],
            header_rows=1,
        )
    )

    story.extend([Spacer(1, 8), Paragraph("Report 2 - Failure Distribution", section_style)])
    story.append(_table([["Category", "Students"], *[[item["label"], item["count"]] for item in analysis["failureDistribution"]]], header_rows=1))

    story.append(PageBreak())
    story.append(Paragraph("Report 3 - Subject Details", section_style))
    for row in rows:
        story.append(Paragraph(row["subject"], section_style))
        story.append(
            _table(
                [["Metric", "Value"], ["Failed Students", len(row["failedStudents"])], ["Absent Students", len(row["absentStudents"])], ["Borderline Students", len(row["borderlineStudents"])]],
                header_rows=1,
            )
        )
        failed_names = ", ".join(item["studentName"] for item in row["failedStudents"][:18]) or "None"
        absent_names = ", ".join(item["studentName"] for item in row["absentStudents"][:18]) or "None"
        borderline_names = ", ".join(item["studentName"] for item in row["borderlineStudents"][:18]) or "None"
        story.append(Paragraph(f"Failed Students: {failed_names}", small_style))
        story.append(Paragraph(f"Absent Students: {absent_names}", small_style))
        story.append(Paragraph(f"Borderline Students: {borderline_names}", small_style))
        story.append(Spacer(1, 6))

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    buffer.seek(0)
    return buffer


def build_excel_export(*, academic_year: str, year: str, semester: str, section: str, exam: str) -> BytesIO:
    analysis, uploads, settings = _load_report_data(
        academic_year=academic_year,
        year=year,
        semester=semester,
        section=section,
        exam=exam,
    )
    rows = _subject_rows(analysis, uploads, settings)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    _excel_header(summary, academic_year, year, semester, section, exam)

    _write_rows(
        summary,
        [
            [],
            [],
            [],
            [],
            ["Metric", "Value"],
            ["Class Strength", analysis["overall"]["classStrength"]],
            ["Overall Pass %", analysis["overall"]["passPercentage"]],
            ["Students Passed", _all_pass_count(analysis)],
            ["Students Failed", max(analysis["overall"]["classStrength"] - _all_pass_count(analysis), 0)],
            ["Upload Progress", f"{len(uploads['uploads'])} / {len(settings.get('subjects', []))}"],
        ],
    )

    report1 = workbook.create_sheet("Report 1")
    _excel_header(report1, academic_year, year, semester, section, exam)
    _write_rows(
        report1,
        [
            [],
            [],
            [],
            [],
            ["Name of the Faculty", "Subject Name", "Subject Code", "Total Strength", "Attended", "Absent", "Passed", "Failed", "% of Pass", "% of Fail"],
            *[
                [
                    row["faculty"],
                    row["subject"],
                    row["subjectCode"],
                    row["strength"],
                    row["attended"],
                    row["absent"],
                    row["passed"],
                    row["failed"],
                    row["passPercentage"],
                    row["failPercentage"],
                ]
                for row in rows
            ],
        ],
    )

    report2 = workbook.create_sheet("Report 2")
    _excel_header(report2, academic_year, year, semester, section, exam)
    _write_rows(report2, [[], [], [], [], ["Failure Distribution", "Students"], *[[item["label"], item["count"]] for item in analysis["failureDistribution"]]])

    report3 = workbook.create_sheet("Report 3")
    _excel_header(report3, academic_year, year, semester, section, exam)
    _write_rows(
        report3,
        [
            [],
            [],
            [],
            [],
            ["Subject Code", "Register Number", "Student Name", "Marks", "Result", "Borderline"],
            *[
                [row["subjectCode"], item["registerNumber"], item["studentName"], item["marks"], item["result"], "Yes" if item["isBorderline"] else "No"]
                for row in rows
                for item in row["marksTable"]
            ],
        ],
    )
    _write_report3_consolidated(report3, analysis, rows)

    for sheet in workbook.worksheets:
        _style_sheet(sheet)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _load_report_data(*, academic_year: str, year: str, semester: str, section: str, exam: str) -> tuple[dict[str, Any], dict[str, Any], dict[str, Any]]:
    settings = get_context_config(academic_year, year, semester, section)
    if not settings.get("subjects") or not settings.get("facultyAssignments"):
        return _empty_analysis(exam), {"uploads": [], "mergedStudentCount": 0, "warnings": []}, settings

    merged = load_merged_data(academic_year=academic_year, year=year, semester=semester, section=section, exam=exam)
    configured_codes = {subject["code"].strip().upper() for subject in settings.get("subjects", [])}
    analysis = analyze_merged_data(_filter_merged_subjects(merged, configured_codes), exam=exam)
    uploads = get_upload_status(academic_year=academic_year, year=year, semester=semester, section=section, exam=exam)
    uploads["uploads"] = [
        upload
        for upload in uploads.get("uploads", [])
        if upload.get("subjectCode", "").strip().upper() in configured_codes
    ]
    if not uploads["uploads"]:
        uploads["mergedStudentCount"] = 0
    return analysis, uploads, settings


def _subject_rows(analysis: dict[str, Any], uploads: dict[str, Any], settings: dict[str, Any]) -> list[dict[str, Any]]:
    upload_by_code = {item["subjectCode"].upper(): item for item in uploads.get("uploads", [])}
    subject_by_code = {item["code"].upper(): item for item in settings.get("subjects", [])}
    faculty_by_code = {item["subjectCode"].upper(): item for item in settings.get("facultyAssignments", [])}
    rows = []
    for subject in analysis["subjects"]:
        code = subject["subjectCode"]
        students = []
        for student in analysis["students"]:
            result = student["subjects"].get(code, {"marks": None, "result": "not_uploaded", "isBorderline": False})
            students.append(
                {
                    "registerNumber": student["registerNumber"],
                    "studentName": student["studentName"],
                    "marks": result["marks"],
                    "result": result["result"],
                    "isBorderline": result["isBorderline"],
                }
            )
        rows.append(
            {
                "faculty": faculty_by_code.get(code.upper(), {}).get("facultyName") or upload_by_code.get(code.upper(), {}).get("facultyName", "Unassigned"),
                "subject": f"{code} - {subject_by_code.get(code.upper(), {}).get('name') or upload_by_code.get(code.upper(), {}).get('subjectName', code)}",
                "subjectName": subject_by_code.get(code.upper(), {}).get("name") or upload_by_code.get(code.upper(), {}).get("subjectName", code),
                "subjectCode": code,
                "strength": subject["classStrength"],
                "attended": subject["studentsAttended"],
                "absent": subject["studentsAbsent"],
                "passed": subject["studentsPassed"],
                "failed": subject["studentsFailed"],
                "passPercentage": subject["passPercentage"],
                "failPercentage": subject["failPercentage"],
                "failedStudents": [item for item in students if item["result"] == "fail"],
                "absentStudents": [item for item in students if item["result"] == "absent"],
                "borderlineStudents": [item for item in students if item["isBorderline"]],
                "marksTable": students,
            }
        )
    return rows


def _filter_merged_subjects(merged: dict[str, Any], configured_codes: set[str]) -> dict[str, Any]:
    students = []
    for student in merged.get("students", []):
        subjects = {
            code: result
            for code, result in student.get("subjects", {}).items()
            if code.strip().upper() in configured_codes
        }
        if subjects:
            students.append({**student, "subjects": subjects})
    return {"students": students, "warnings": merged.get("warnings", [])}


def _all_pass_count(analysis: dict[str, Any]) -> int:
    for item in analysis["failureDistribution"]:
        if item["label"] == "All Pass":
            return int(item["count"])
    return 0


def _empty_analysis(exam: str) -> dict[str, Any]:
    exam_type = detect_exam_type(exam)
    rule_by_type = {
        "UT": {"total": 25, "passMark": 15, "borderlineMin": 15, "borderlineMax": 17},
        "CAT": {"total": 50, "passMark": 33, "borderlineMin": 33, "borderlineMax": 35},
        "END_SEMESTER": {"total": None, "passMark": None, "borderlineMin": None, "borderlineMax": None},
    }
    zero_metrics = {
        "classStrength": 0,
        "studentsAttended": 0,
        "studentsAbsent": 0,
        "studentsPassed": 0,
        "studentsFailed": 0,
        "passPercentage": 0,
        "failPercentage": 0,
        "averageMarks": None,
        "highestMarks": None,
        "lowestMarks": None,
    }
    return {
        "examType": exam_type,
        "rule": rule_by_type[exam_type],
        "overall": zero_metrics,
        "subjects": [],
        "students": [],
        "failureDistribution": [
            {"label": "All Pass", "count": 0},
            {"label": "One Failure", "count": 0},
            {"label": "Two Failures", "count": 0},
            {"label": "Three Failures", "count": 0},
            {"label": "Four Failures", "count": 0},
            {"label": "More than Four Failures", "count": 0},
        ],
        "warnings": ["No configured subjects or faculty assignments for the selected context."],
    }


def _write_report3_consolidated(sheet: Any, analysis: dict[str, Any], rows: list[dict[str, Any]]) -> None:
    sheet.append([])
    sheet.append([])
    sheet.append(["Consolidated Student Marks"])
    subject_headers = [f"Subject {index + 1} Marks - {row['subjectName']}" for index, row in enumerate(rows)]
    sheet.append(["Register Number", "Student Name", *subject_headers])

    for student in analysis["students"]:
        marks_by_subject = []
        for row in rows:
            result = student["subjects"].get(row["subjectCode"], {})
            marks_by_subject.append(result.get("marks", ""))
        sheet.append([student["registerNumber"], student["studentName"], *marks_by_subject])

    total_students = analysis["overall"]["classStrength"]
    passed = _all_pass_count(analysis)
    failed = max(total_students - passed, 0)
    borderline = sum(
        1
        for student in analysis["students"]
        if any(result.get("isBorderline") for result in student.get("subjects", {}).values())
    )
    absent = sum(
        1
        for student in analysis["students"]
        if any(result.get("result") == "absent" for result in student.get("subjects", {}).values())
    )

    sheet.append([])
    sheet.append(["Consolidated Summary", "Value"])
    sheet.append(["Total Students", total_students])
    sheet.append(["Passed", passed])
    sheet.append(["Failed", failed])
    sheet.append(["Borderline", borderline])
    sheet.append(["Absent", absent])
    sheet.append(["Overall Pass %", analysis["overall"]["passPercentage"]])
    sheet.append(["Overall Fail %", analysis["overall"]["failPercentage"]])


def _table(data: list[list[Any]], *, header_rows: int) -> Table:
    table = Table(data, repeatRows=header_rows, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, header_rows - 1), colors.HexColor("#fff9c4")),
                ("TEXTCOLOR", (0, 0), (-1, header_rows - 1), colors.black),
                ("FONTNAME", (0, 0), (-1, header_rows - 1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
                ("ROWBACKGROUNDS", (0, header_rows), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    return table


def _footer(canvas: Any, doc: Any) -> None:
    canvas.saveState()
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.HexColor("#64748b"))
    canvas.drawString(doc.leftMargin, 0.22 * inch, f"{COLLEGE_NAME} - {DEPARTMENT_NAME}")
    canvas.drawRightString(doc.pagesize[0] - doc.rightMargin, 0.22 * inch, f"Page {doc.page}")
    canvas.restoreState()


def _write_rows(sheet: Any, rows: list[list[Any]]) -> None:
    for row in rows:
        sheet.append(row)


def _style_sheet(sheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="FFF9C4")
    header_font = Font(color="000000", bold=True)
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    header_labels = {"Metric", "Name of the Faculty", "Failure Distribution", "Subject Code", "Register Number", "Consolidated Summary", "Consolidated Student Marks"}
    for row in sheet.iter_rows():
        first_value = str(row[0].value or "")
        if first_value in header_labels:
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font
    for column in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = min(max(max_length + 2, 12), 32)


def _pdf_header(academic_year: str, year: str, semester: str, section: str, exam: str, title_style: ParagraphStyle, small_style: ParagraphStyle) -> Table:
    logo = PdfImage(str(LOGO_PATH), width=0.55 * inch, height=0.55 * inch) if LOGO_PATH.exists() else ""
    heading = [
        Paragraph(f"<b>{COLLEGE_NAME}</b>", title_style),
        Paragraph("Bharathi Salai<br/>Ramapuram<br/>Chennai - 600089", small_style),
        Paragraph(f"<b>{DEPARTMENT_NAME.upper()}</b>", small_style),
        Paragraph(f"<b>{exam} ANALYSIS</b>", small_style),
        Paragraph(f"Academic Year: {academic_year} | Year: {year} | Semester: {semester} | Section: {section} | Exam: {exam}", small_style),
    ]
    table = Table([[logo, heading]], colWidths=[0.85 * inch, 9.2 * inch])
    table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (0, 0), "CENTER"),
                ("ALIGN", (1, 0), (1, 0), "CENTER"),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
                ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    return table


def _excel_header(sheet: Any, academic_year: str, year: str, semester: str, section: str, exam: str) -> None:
    sheet.row_dimensions[1].height = 24
    sheet.row_dimensions[2].height = 20
    sheet.row_dimensions[3].height = 20
    sheet.row_dimensions[4].height = 20
    sheet.row_dimensions[5].height = 20
    if LOGO_PATH.exists():
        try:
            image = ExcelImage(str(LOGO_PATH))
            image.width = 48
            image.height = 48
            sheet.add_image(image, "A1")
        except Exception:
            pass
    sheet.merge_cells("B1:J1")
    sheet.merge_cells("B2:J2")
    sheet.merge_cells("B3:J3")
    sheet.merge_cells("A4:J4")
    sheet.merge_cells("A5:J5")
    sheet["B1"] = COLLEGE_NAME
    sheet["B2"] = "Bharathi Salai, Ramapuram, Chennai - 600089"
    sheet["B3"] = f"{exam} ANALYSIS"
    sheet["A4"] = f"DEPARTMENT : {DEPARTMENT_NAME.upper()}"
    sheet["A5"] = f"ACADEMIC YEAR : {academic_year}    YEAR : {year}    SEMESTER : {semester}    SECTION : {section}    EXAM : {exam}"
    for cell in ["B1", "B2", "B3", "A4", "A5"]:
        sheet[cell].font = Font(bold=True)
        sheet[cell].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
COLLEGE_NAME = "Easwari Engineering College"
COLLEGE_ADDRESS = "Bharathi Salai, Ramapuram, Chennai - 600089."
DEPARTMENT_NAME = "Department of Artificial Intelligence and Data Science"
LOGO_PATH = STORAGE_DIR / "assets" / "college_logo.png"
